"""
Spreadsheet Analyzer Tool: Reads and analyzes Excel (.xlsx) files using openpyxl.
"""

import logging
from pathlib import Path
from typing import List

from openpyxl import load_workbook

from backend.tools.path_safety import safe_resolve_output_path

logger = logging.getLogger(__name__)

# Base directory for reading spreadsheet files
BASE_DIR = Path(__file__).parent.parent.parent / "workspace" / "sandbox_files"


def read_sheet(filename: str, cell_range: str = "A1:D50") -> List[List[str]]:
    """
    Read data from an Excel spreadsheet within a specified cell range.

    Args:
        filename: Name of the spreadsheet file (resolved against sandbox_files/).
        cell_range: Excel-style cell range (default "A1:D50").

    Returns:
        A 2D list of strings containing the cell values.
    """
    # Resolve path with containment check (prevents path traversal)
    file_path = safe_resolve_output_path(filename, BASE_DIR)

    if not file_path.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {file_path}")

    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        ws = wb.active

        # Parse the cell range (e.g. "A1:D50")
        cells = ws[cell_range]

        data = []
        for row in cells:
            row_data = []
            for cell in row:
                value = cell.value
                row_data.append(str(value) if value is not None else "")
            data.append(row_data)

        wb.close()

        logger.info(f"Read {len(data)} rows from {filename} range {cell_range}")
        return data

    except Exception as e:
        logger.error(f"Spreadsheet read failed: {e}")
        raise
